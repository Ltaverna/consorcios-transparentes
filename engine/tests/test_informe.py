"""El informe (Excel y HTML) se genera desde el modelo, sin datos manuales."""
import os

import pytest

from ct.informe import informe_excel, informe_html
from ct.redconar import parse_text
from ct.rules import Config, evaluar

FIX = os.path.join(os.path.dirname(__file__), "fixtures")


def _liq(name):
    return parse_text(open(os.path.join(FIX, name), encoding="utf-8").read())


def test_html_autocontenido(tmp_path):
    liq = _liq("redconar_202608.txt"); prev = _liq("redconar_202607.txt")
    hs = evaluar(liq, prev, Config())
    out = informe_html(liq, hs, str(tmp_path / "i.html"), prev, None, marca="Prueba")
    h = open(out, encoding="utf-8").read()
    assert h.startswith("<!doctype html>") and 'name="viewport"' in h
    assert "Prueba" in h and "Agosto 2026" in h and "Julio 2026" in h
    assert "Hallazgos" in h and "Cuadre de la liquidación" in h
    assert h.count('class="h"') == len(hs)
    assert "$ 29.876.923,16" in h


def test_excel_hojas(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    liq = _liq("redconar_202608.txt"); prev = _liq("redconar_202607.txt")
    hs = evaluar(liq, prev, Config())
    out = informe_excel(liq, hs, str(tmp_path / "i.xlsx"), prev)
    wb = load_workbook(out)
    assert [ws.title for ws in wb.worksheets] == ["Resumen", "Hallazgos", "Gastos por rubro", "Gastos detalle", "Proveedores", "Deudores", "Estado de cuentas", "Evolución"]
    assert wb["Hallazgos"].max_row == len(hs) + 1
    assert wb["Estado de cuentas"].max_row == 117
    assert wb["Gastos detalle"].max_row == 44


def test_html_liquidacion_2024_sin_unidades(tmp_path):
    liq = _liq("redconar_202407.txt")
    hs = evaluar(liq, None, Config())
    out = informe_html(liq, hs, str(tmp_path / "i.html"))
    assert "Deudores" in open(out, encoding="utf-8").read()
