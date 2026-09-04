"""Generación de informes (Excel y HTML) a partir del análisis del motor.

informe_excel(): requiere openpyxl (pip install openpyxl).
informe_html(): sin dependencias; página autocontenida, apta para celular, con marca opcional.
"""
from __future__ import annotations
import html
import json
from collections import OrderedDict
from datetime import date
from typing import Optional

from .model import Liquidacion
from .rules import Hallazgo

SEV_ORDER = {"CRÍTICO": 0, "ALTO": 1, "MEDIO": 2, "BAJO": 3}


def _money(v: float) -> str:
    return ("-" if v < 0 else "") + "$ " + f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _pct(v: float) -> str:
    return f"{v * 100:.1f} %".replace(".", ",")


def _kpis(liq: Liquidacion, prev: Optional[Liquidacion]) -> list[dict]:
    e = liq.estado
    cobrado = e.ing_termino + e.ing_adeudadas + e.ing_intereses + e.ing_adelantadas
    caja = next((c for c in liq.cuentas if "CAJA" in c.nombre.upper()), None)
    deuda = sum(u.deuda for u in liq.unidades if u.deuda > 0) or (liq.total_deudores or 0)
    pend = abs(liq.patrimonial.facturas_pend)
    efectivo = sum(g.importe for g in liq.gastos if g.en_efectivo)
    unid = sum(g.importe for g in liq.gastos if "UNIDADES" in g.categoria.upper())
    tot_pr = liq.prorrateo_total.get("_total_mes") or 0

    def pv(fn):
        try:
            return fn(prev) if prev else None
        except Exception:
            return None

    rows = [
        ("Gastos del mes", liq.suma_gastos, pv(lambda p: p.suma_gastos), "", False),
        ("Importe prorrateado a cobrar", tot_pr, pv(lambda p: p.prorrateo_total.get("_total_mes") or 0), "Diferencia con el gasto: " + _money(tot_pr - liq.suma_gastos) if tot_pr else "", False),
        ("Expensas cobradas en el mes", cobrado, pv(lambda p: p.estado.ing_termino + p.estado.ing_adeudadas + p.estado.ing_intereses + p.estado.ing_adelantadas), "", True),
        ("Disponibilidades al cierre", e.saldo_cierre, pv(lambda p: p.estado.saldo_cierre), "", True),
        ("   en efectivo (caja)", caja.saldo_cierre if caja else 0, pv(lambda p: next((c.saldo_cierre for c in p.cuentas if "CAJA" in c.nombre.upper()), 0)), _pct(caja.saldo_cierre / e.saldo_cierre) + " de la liquidez" if caja and e.saldo_cierre else "", False),
        ("Facturas pendientes de pago", pend, pv(lambda p: abs(p.patrimonial.facturas_pend)), "", False),
        ("Disponibilidades menos pendientes", e.saldo_cierre - pend, pv(lambda p: p.estado.saldo_cierre - abs(p.patrimonial.facturas_pend)), "", True),
        ("Deuda de propietarios", deuda, pv(lambda p: sum(u.deuda for u in p.unidades if u.deuda > 0) or (p.total_deudores or 0)), f"{len([u for u in liq.unidades if u.deuda > 0]) or len(liq.deudores)} unidades", False),
        ("Pagado en efectivo a proveedores", efectivo, pv(lambda p: sum(g.importe for g in p.gastos if g.en_efectivo)), "", False),
        ("Trabajos en unidades privadas", unid, pv(lambda p: sum(g.importe for g in p.gastos if "UNIDADES" in g.categoria.upper())), _pct(unid / liq.suma_gastos) + " del gasto" if liq.suma_gastos else "", False),
    ]
    return [dict(label=l, actual=a, anterior=b, nota=n, bueno_si_sube=up) for l, a, b, n, up in rows]


# ------------------------------------------------------------------ EXCEL
def informe_excel(liq: Liquidacion, hallazgos: list[Hallazgo], out: str, prev: Optional[Liquidacion] = None, docs: Optional[list] = None, marca: str = "") -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    MONEY = '#,##0.00;[Red]-#,##0.00'; PCT = '0.0%'
    HDR = PatternFill("solid", fgColor="1F2A44"); HF = Font(bold=True, color="FFFFFF", size=10)
    T = Font(bold=True, size=14, color="17253A"); S = Font(bold=True, size=11, color="17253A"); N = Font(italic=True, size=9, color="5A6270"); B = Font(bold=True, size=10); R = Font(size=10)
    FILL = {"CRÍTICO": "F8D0D0", "ALTO": "FBE3CC", "MEDIO": "FFF2C2", "BAJO": "E3EEF9"}
    thin = Border(bottom=Side(style="thin", color="C9CFD9")); wrap = Alignment(wrap_text=True, vertical="top")
    wb = Workbook()

    def sheet(name, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = name[:31]; ws.sheet_view.showGridLines = False
        return ws

    def header(ws, r, cols, widths=None):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=r, column=j, value=c); cell.fill = HDR; cell.font = HF; cell.alignment = Alignment(wrap_text=True, vertical="center")
        if widths:
            for j, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(j)].width = w
        ws.row_dimensions[r].height = 28
        return r + 1

    def put(ws, r, vals, fmts=None, bold=False, fill=None, wrapc=False):
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v); c.font = B if bold else R; c.border = thin
            if fmts and j - 1 < len(fmts) and fmts[j - 1]:
                c.number_format = fmts[j - 1]
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            if wrapc:
                c.alignment = wrap
        return r + 1

    titulo = f"{liq.consorcio} · {liq.periodo}"
    # ---- Resumen
    ws = sheet("Resumen", first=True)
    ws.cell(row=1, column=1, value=(marca + " · " if marca else "") + "Análisis de la liquidación de expensas").font = T
    ws.cell(row=2, column=1, value=f"{titulo} · Administración: {liq.administracion} · Generado el {date.today():%d-%m-%Y} con Consorcio Transparente").font = N
    for col, w in zip("ABCDE", (44, 20, 20, 12, 60)):
        ws.column_dimensions[col].width = w
    r = 4; ws.cell(row=r, column=1, value="Indicadores").font = S; r += 1
    r = header(ws, r, ["Indicador", liq.periodo, prev.periodo if prev else "Mes anterior", "Variación", "Nota"])
    for k in _kpis(liq, prev):
        put(ws, r, [k["label"], k["actual"], k["anterior"], None, k["nota"]], [None, MONEY, MONEY, PCT, None])
        if k["anterior"]:
            ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/C{r}-1,\"\")").number_format = PCT
        r += 1
    r += 1; ws.cell(row=r, column=1, value="Cuadre de la liquidación").font = S; r += 1
    r = header(ws, r, ["Verificación", "Esperado", "Obtenido", "Diferencia", "Resultado"])
    for c in liq.checks:
        r = put(ws, r, [c.nombre, c.esperado, c.obtenido, c.diff, "OK" if c.ok else "FALLA"], [None, MONEY, MONEY, MONEY, None], fill=None if c.ok else FILL["CRÍTICO"])
    r += 1; ws.cell(row=r, column=1, value="Hallazgos por severidad").font = S; r += 1
    for sev in SEV_ORDER:
        n = sum(1 for h in hallazgos if h.severidad == sev)
        if n:
            r = put(ws, r, [sev, n], fill=FILL[sev])
    # ---- Hallazgos
    ws = sheet("Hallazgos")
    r = header(ws, 1, ["#", "Severidad", "Área", "Hallazgo", "Evidencia", "Monto", "Qué pedir", "Regla", "Refs"], [4, 10, 24, 48, 80, 16, 50, 14, 10])
    for i, h in enumerate(sorted(hallazgos, key=lambda h: (SEV_ORDER.get(h.severidad, 9), -abs(h.monto))), 1):
        put(ws, r, [i, h.severidad, h.area, h.titulo, h.evidencia, h.monto, h.recomendacion, h.regla, ", ".join(h.refs)], [None] * 5 + [MONEY], wrapc=True)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=FILL.get(h.severidad, "FFFFFF")); ws.cell(row=r, column=4).font = B
        ws.row_dimensions[r].height = 90; r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{r - 1}"
    # ---- Rubros
    ws = sheet("Gastos por rubro")
    r = header(ws, 1, ["Rubro", liq.periodo, "% del mes", prev.periodo if prev else "Mes anterior", "Variación"], [36, 18, 10, 18, 11])
    cats = liq.por_categoria(); pc = prev.por_categoria() if prev else {}
    for cat, v in sorted(cats.items(), key=lambda kv: -kv[1]):
        put(ws, r, [cat, v, v / liq.suma_gastos if liq.suma_gastos else 0, pc.get(cat), None], [None, MONEY, PCT, MONEY, PCT])
        if pc.get(cat):
            ws.cell(row=r, column=5, value=f"=IFERROR(B{r}/D{r}-1,\"\")").number_format = PCT
        r += 1
    put(ws, r, ["Total", liq.suma_gastos, 1, prev.suma_gastos if prev else None], [None, MONEY, PCT, MONEY], bold=True)
    # ---- Detalle
    ws = sheet("Gastos detalle")
    r = header(ws, 1, ["#", "Rubro", "Proveedor", "Concepto", "Clase", "Importe", "% del mes", "Fecha factura", "N° factura", "Importe factura", "Fecha pago", "Caja", "Forma", "Días factura→pago", "Período"], [4, 24, 30, 70, 5, 16, 8, 12, 18, 16, 12, 8, 18, 9, 14])
    for g in liq.gastos:
        p0 = g.pagos[0] if g.pagos else None
        put(ws, r, [g.n, g.categoria, g.proveedor, g.concepto[:400], g.columna, g.importe, g.importe / liq.suma_gastos if liq.suma_gastos else 0, g.factura_fecha, g.factura_nro, g.factura_importe, g.fecha_pago, p0.caja if p0 else "", p0.forma if p0 else "", g.dias_factura_pago, g.periodo],
            [None] * 5 + [MONEY, PCT, "DD-MM-YYYY", None, MONEY, "DD-MM-YYYY"], wrapc=True, fill=FILL["ALTO"] if g.en_efectivo else None)
        ws.row_dimensions[r].height = 32; r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:O{r - 1}"
    # ---- Proveedores
    ws = sheet("Proveedores")
    r = header(ws, 1, ["#", "Proveedor", liq.periodo, prev.periodo if prev else "Mes anterior", "% del mes", "% acumulado", "Líneas"], [4, 40, 18, 18, 10, 11, 8])
    pp = prev.por_proveedor() if prev else {}; acc = 0.0
    for i, (p, v) in enumerate(liq.por_proveedor().items(), 1):
        acc += v
        r = put(ws, r, [i, p, v, pp.get(p), v / liq.suma_gastos if liq.suma_gastos else 0, acc / liq.suma_gastos if liq.suma_gastos else 0, sum(1 for g in liq.gastos if g.proveedor == p)], [None, None, MONEY, MONEY, PCT, PCT, "0"])
    # ---- Deudores y unidades
    if liq.unidades:
        ws = sheet("Deudores")
        r = header(ws, 1, ["UF", "Piso-Depto", "Propietario", "Deuda", "Pagó en el mes", "Interés", "Tasa", "Expensa del mes", "Meses de deuda", "% de la deuda"], [6, 10, 30, 16, 16, 14, 8, 16, 10, 10])
        deud = sorted([u for u in liq.unidades if u.deuda > 0], key=lambda u: -u.deuda); tot = sum(u.deuda for u in deud) or 1
        for u in deud:
            r = put(ws, r, [u.uf, u.piso_depto, u.propietario, u.deuda, u.pagos, u.interes, (u.interes / u.deuda) if u.deuda else 0, u.total_mes, (u.deuda / u.total_mes) if u.total_mes else 0, u.deuda / tot], [None, None, None, MONEY, MONEY, MONEY, PCT, MONEY, "0.0", PCT], fill=FILL["CRÍTICO"] if u.pagos == 0 else None)
        put(ws, r, ["Total", "", "", sum(u.deuda for u in deud), sum(u.pagos for u in deud), sum(u.interes for u in deud)], [None, None, None, MONEY, MONEY, MONEY], bold=True)
        ws = sheet("Estado de cuentas")
        classes = [k for k in liq.prorrateo_total if not k.startswith("_")]
        cols = ["UF", "Piso-Depto", "Propietario", "Tipo", "Saldo anterior", "Pagos", "Créd./Déb.", "Deuda", "Interés"] + [x for c in classes for x in (f"% {c}", f"Expensa {c}")] + ["Expensa del mes", "G. particulares", "Redondeo", "A pagar"]
        r = header(ws, 1, cols, [6, 10, 30, 12] + [15] * 5 + [7, 14] * len(classes) + [15, 12, 9, 15])
        for u in liq.unidades:
            vals = [u.uf, u.piso_depto, u.propietario, u.tipo, u.saldo_ant, u.pagos, u.cred_deb, u.deuda, u.interes] + [x for c in classes for x in (u.pcts.get(c, 0) / 100, u.expensas.get(c, 0))] + [u.total_mes, u.gastos_part, u.redondeo, u.a_pagar]
            fm = [None] * 4 + [MONEY] * 5 + ["0.00%", MONEY] * len(classes) + [MONEY] * 4
            r = put(ws, r, vals, fm, fill=FILL["CRÍTICO"] if (u.deuda > 0 and u.pagos == 0) else (FILL["ALTO"] if u.deuda > 0 else None))
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r - 1}"
    elif liq.deudores:
        ws = sheet("Deudores")
        r = header(ws, 1, ["UF", "Piso-Depto", "Propietario", "Deuda"], [6, 10, 30, 16])
        for d in liq.deudores:
            r = put(ws, r, [d.uf, d.piso_depto, d.propietario, d.deuda], [None, None, None, MONEY])
    # ---- Evolución
    if liq.evolucion:
        ws = sheet("Evolución")
        r = header(ws, 1, ["Mes", "Importe a cobrar", "Gastos del mes", "Expensas cobradas", "Cobrado / prorrateado", "Gastos / prorrateado"], [16, 18, 18, 18, 14, 14])
        for m in liq.evolucion:
            r = put(ws, r, [m.mes, m.a_cobrar, m.gastos, m.cobrado, (m.cobrado / m.a_cobrar) if m.a_cobrar else 0, (m.gastos / m.a_cobrar) if m.a_cobrar else 0], [None, MONEY, MONEY, MONEY, PCT, PCT])
    # ---- Comprobantes
    if docs:
        ws = sheet("Comprobantes")
        r = header(ws, 1, ["Gasto #", "Archivo", "Tipo", "Emisor CUIT", "Receptor", "Receptor CUIT", "Tipo fact.", "N° factura", "Fecha", "Importe", "Destinatario", "Destinatario CUIT", "Motivo", "Operación", "Notas"], [7, 60, 9, 13, 30, 13, 8, 16, 11, 14, 30, 13, 26, 14, 50])
        for d in docs:
            r = put(ws, r, [d.gasto_n, d.archivo, d.tipo, d.emisor_cuit, d.receptor, d.receptor_cuit, d.factura_tipo, d.factura_nro, d.fecha, d.importe, d.destinatario, d.destinatario_cuit, d.motivo, d.operacion, "; ".join(d.notas)], [None] * 8 + ["DD-MM-YYYY", MONEY], wrapc=True)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:O{r - 1}"
    wb.save(out)
    return out


# ------------------------------------------------------------------ HTML
CSS = """
:root{--bg:#f2f3f5;--surface:#fff;--surface-2:#f7f8fa;--ink:#1b2536;--ink-2:#414a58;--muted:#5f6875;--hair:#d5dae2;--hair-2:#e8ebf0;--accent:#2a5db0;--s1:#2a78d6;--s2:#eb6834;--crit:#d03b3b;--ser:#ec835a;--warn:#c98500;--info:#2a78d6;--crit-s:#fbe4e4;--ser-s:#fdeadf;--warn-s:#fff3cf;--info-s:#e3eefb;--good:#0b7d0b;color-scheme:light}
@media (prefers-color-scheme:dark){:root:not([data-theme="light"]){--bg:#0f1318;--surface:#171c24;--surface-2:#1d232c;--ink:#eef1f5;--ink-2:#c3cad4;--muted:#8e98a6;--hair:#2a313c;--hair-2:#222932;--accent:#6f9ee8;--s1:#3987e5;--s2:#d95926;--crit:#e66767;--ser:#ec835a;--warn:#fab219;--info:#6f9ee8;--crit-s:#3a1d1d;--ser-s:#3b261a;--warn-s:#3a3113;--info-s:#1c2a42;--good:#7fe07f;color-scheme:dark}}
:root[data-theme="dark"]{--bg:#0f1318;--surface:#171c24;--surface-2:#1d232c;--ink:#eef1f5;--ink-2:#c3cad4;--muted:#8e98a6;--hair:#2a313c;--hair-2:#222932;--accent:#6f9ee8;--s1:#3987e5;--s2:#d95926;--crit:#e66767;--ser:#ec835a;--warn:#fab219;--info:#6f9ee8;--crit-s:#3a1d1d;--ser-s:#3b261a;--warn-s:#3a3113;--info-s:#1c2a42;--good:#7fe07f;color-scheme:dark}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font-family:"IBM Plex Sans",system-ui,-apple-system,"Segoe UI",sans-serif;font-size:15px;line-height:1.5}
h1,h2,h3{font-family:"Source Serif 4",Georgia,serif;font-weight:600;letter-spacing:-.01em;margin:0;text-wrap:balance}h1{font-size:clamp(26px,4vw,38px);line-height:1.1}h2{font-size:23px}h3{font-size:16px}
p{margin:0}.wrap{max-width:1100px;margin:0 auto;padding:0 18px}header{background:var(--surface);border-bottom:1px solid var(--hair)}header .wrap{padding:32px 18px 24px;display:grid;gap:10px}
.eyebrow{font-size:12px;font-weight:600;letter-spacing:.08em;text-transform:uppercase;color:var(--muted)}.lede{color:var(--ink-2);max-width:70ch}
section{padding:32px 0 4px}section .wrap{display:grid;gap:14px}.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px}
.tile{background:var(--surface);border:1px solid var(--hair);border-radius:10px;padding:12px 14px;display:grid;gap:3px}.tile .lab{font-size:12.5px;color:var(--muted);font-weight:500}.tile .val{font-size:22px;font-weight:600;font-variant-numeric:tabular-nums}.tile .sub{font-size:12.5px;color:var(--ink-2)}
.tile.warn{box-shadow:inset 3px 0 0 var(--crit)}.d{font-size:12.5px;font-weight:500}.d.up{color:var(--crit)}.d.down{color:var(--good)}.d.flat{color:var(--muted)}
.h{background:var(--surface);border:1px solid var(--hair);border-radius:10px;border-left:4px solid var(--c);padding:12px 14px;display:grid;gap:6px}.h .sev{display:inline-block;font-size:11px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;padding:2px 8px;border-radius:5px;background:var(--cs)}
.h .ttl{font-weight:600}.h .ev{font-size:14px;color:var(--ink-2)}.h .rec{font-size:14px;background:var(--surface-2);border-radius:8px;padding:8px 10px}.h[data-sev="CRÍTICO"]{--c:var(--crit);--cs:var(--crit-s)}.h[data-sev="ALTO"]{--c:var(--ser);--cs:var(--ser-s)}.h[data-sev="MEDIO"]{--c:var(--warn);--cs:var(--warn-s)}.h[data-sev="BAJO"]{--c:var(--info);--cs:var(--info-s)}
.bars{display:grid;gap:6px}.bar{display:grid;grid-template-columns:minmax(0,200px) minmax(0,1fr) auto;gap:10px;align-items:center;font-size:14px}.bar .lab{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.bar .track{height:10px;background:var(--hair-2);border-radius:5px;overflow:hidden}.bar .track i{display:block;height:100%;background:var(--s1);border-radius:5px}.bar .v{font-variant-numeric:tabular-nums;white-space:nowrap;font-weight:600}
.tw{overflow-x:auto;border:1px solid var(--hair);border-radius:10px;background:var(--surface)}table{border-collapse:collapse;width:100%;font-size:13.5px}th{text-align:left;background:var(--surface-2);font-size:12px;color:var(--ink-2);padding:8px 10px;border-bottom:1px solid var(--hair);white-space:nowrap}td{padding:7px 10px;border-bottom:1px solid var(--hair-2);vertical-align:top}td.r,th.r{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.ok{color:var(--good);font-weight:600}.fail{color:var(--crit);font-weight:600}.note{font-size:13px;color:var(--muted)}footer{padding:30px 0 50px;color:var(--muted);font-size:13px}
@media (max-width:640px){.bar{grid-template-columns:minmax(0,110px) minmax(0,1fr) auto}.tile .val{font-size:19px}}
"""


def informe_html(liq: Liquidacion, hallazgos: list[Hallazgo], out: str, prev: Optional[Liquidacion] = None, docs: Optional[list] = None, marca: str = "") -> str:
    E = html.escape
    kp = _kpis(liq, prev)
    hs = sorted(hallazgos, key=lambda h: (SEV_ORDER.get(h.severidad, 9), -abs(h.monto)))
    cats = sorted(liq.por_categoria().items(), key=lambda kv: -kv[1]); mx = cats[0][1] if cats else 1
    provs = list(liq.por_proveedor().items())[:12]; mp = provs[0][1] if provs else 1
    deud = sorted([u for u in liq.unidades if u.deuda > 0], key=lambda u: -u.deuda)
    bad = [c for c in liq.checks if not c.ok]
    nsev = {s: sum(1 for h in hs if h.severidad == s) for s in SEV_ORDER}

    def tile(k):
        a, b = k["actual"], k["anterior"]
        d = ""
        if b:
            v = a / b - 1; peor = (v > 0) != bool(k["bueno_si_sube"]); cls = "flat" if abs(v) < 0.005 else ("up" if peor else "down")
            d = f'<div class="d {cls}">{"+" if v > 0 else ""}{_pct(v)} vs. {E(prev.periodo)}</div>'
        warn = " warn" if k["label"].strip() in ("en efectivo (caja)", "Facturas pendientes de pago", "Deuda de propietarios", "Pagado en efectivo a proveedores", "Trabajos en unidades privadas") and a > 0 else ""
        return f'<div class="tile{warn}"><div class="lab">{E(k["label"].strip())}</div><div class="val">{_money(a)}</div>{d}<div class="sub">{E(k["nota"])}</div></div>'

    parts = [f"""<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Expensas {E(liq.periodo)}</title>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Source+Serif+4:opsz,wght@8..60,600&family=IBM+Plex+Sans:wght@400;500;600&display=swap"><style>{CSS}</style>
<header><div class="wrap"><div class="eyebrow">{E(marca + " · " if marca else "")}{E(liq.consorcio)} · CUIT {E(liq.cuit_consorcio)}</div><h1>Liquidación de expensas de {E(liq.periodo)}, verificada</h1>
<p class="lede">Análisis automático de la liquidación emitida por {E(liq.administracion)}{" contrastada con " + E(prev.periodo) if prev else ""}{" y de " + str(len(docs)) + " comprobantes del portal" if docs else ""}. Cuadre: {len(liq.checks) - len(bad)} de {len(liq.checks)} verificaciones correctas. Hallazgos: {len(hs)} ({", ".join(f"{n} {s.lower()}" for s, n in nsev.items() if n)}).</p></div></header>
<section><div class="wrap"><h2>Resumen del mes</h2><div class="tiles">{"".join(tile(k) for k in kp)}</div></div></section>
<section><div class="wrap"><h2>Hallazgos</h2><p class="lede">Hechos documentados con su evidencia y qué pedir a la administración. No son conclusiones: son preguntas con respaldo.</p>"""]
    for h in hs:
        parts.append(f'<div class="h" data-sev="{E(h.severidad)}"><span class="sev">{E(h.severidad)} · {E(h.area)}</span><div class="ttl">{E(h.titulo)}</div>' + (f'<div class="ev">{E(h.evidencia)}</div>' if h.evidencia else "") + (f'<div class="note">Monto involucrado: {_money(h.monto)}</div>' if h.monto else "") + (f'<div class="rec"><b>Qué pedir.</b> {E(h.recomendacion)}</div>' if h.recomendacion else "") + "</div>")
    parts.append("</div></section>")
    parts.append('<section><div class="wrap"><h2>Gastos por rubro</h2><div class="bars">' + "".join(f'<div class="bar"><span class="lab" title="{E(c)}">{E(c.title())}</span><span class="track"><i style="width:{v / mx * 100:.1f}%"></i></span><span class="v">{_money(v)}</span></div>' for c, v in cats) + f'</div><p class="note">Total {_money(liq.suma_gastos)}</p></div></section>')
    parts.append('<section><div class="wrap"><h2>Proveedores que más cobran</h2><div class="bars">' + "".join(f'<div class="bar"><span class="lab" title="{E(p)}">{E(p)}</span><span class="track"><i style="width:{v / mp * 100:.1f}%"></i></span><span class="v">{_money(v)}</span></div>' for p, v in provs) + "</div></div></section>")
    if deud:
        tot = sum(u.deuda for u in deud)
        parts.append(f'<section><div class="wrap"><h2>Deudores</h2><p class="lede">{len(deud)} unidades deben {_money(tot)}.</p><div class="tw"><table><tr><th>Unidad</th><th>Propietario</th><th class="r">Deuda</th><th class="r">Pagó</th><th class="r">Interés</th><th class="r">Meses</th><th class="r">% deuda</th></tr>' + "".join(f'<tr><td>{E(u.piso_depto)}</td><td>{E(u.propietario)}</td><td class="r">{_money(u.deuda)}</td><td class="r">{_money(u.pagos) if u.pagos else "—"}</td><td class="r">{_money(u.interes)}</td><td class="r">{(u.deuda / u.total_mes) if u.total_mes else 0:.1f}</td><td class="r">{_pct(u.deuda / tot)}</td></tr>' for u in deud) + "</table></div></div></section>")
    elif liq.deudores:
        parts.append('<section><div class="wrap"><h2>Deudores</h2><div class="tw"><table><tr><th>Unidad</th><th>Propietario</th><th class="r">Deuda</th></tr>' + "".join(f'<tr><td>{E(d.piso_depto)}</td><td>{E(d.propietario)}</td><td class="r">{_money(d.deuda)}</td></tr>' for d in liq.deudores) + "</table></div></div></section>")
    parts.append('<section><div class="wrap"><h2>Detalle de gastos</h2><div class="tw"><table><tr><th>#</th><th>Rubro</th><th>Proveedor</th><th>Concepto</th><th>Clase</th><th class="r">Importe</th><th>Factura</th><th>Pago</th></tr>' + "".join(f'<tr><td>{g.n}</td><td>{E(g.categoria.title())}</td><td>{E(g.proveedor)}</td><td>{E(g.concepto[:140])}</td><td>{E(g.columna)}</td><td class="r">{_money(g.importe)}</td><td>{E(g.factura_nro or "")}{" · " + g.factura_fecha.strftime("%d-%m") if g.factura_fecha else ""}</td><td>{g.fecha_pago.strftime("%d-%m") if g.fecha_pago else ""}{" · " + E(g.pagos[0].forma) if g.pagos else ""}{" · <b>EFECTIVO</b>" if g.en_efectivo else ""}</td></tr>' for g in liq.gastos) + "</table></div></div></section>")
    if docs:
        parts.append('<section><div class="wrap"><h2>Comprobantes leídos</h2><div class="tw"><table><tr><th>Gasto</th><th>Archivo</th><th>Tipo</th><th>Emisor CUIT</th><th>Receptor</th><th>Destinatario (CUIT)</th><th>Fecha</th><th class="r">Importe</th><th>Notas</th></tr>' + "".join(f'<tr><td>{d.gasto_n or ""}</td><td>{E(d.archivo[:60])}</td><td>{E(d.tipo)}</td><td>{E(d.emisor_cuit or "")}</td><td>{E((d.receptor or "")[:40])}</td><td>{E((d.destinatario or "")[:30])}{" (" + d.destinatario_cuit + ")" if d.destinatario_cuit else ""}</td><td>{d.fecha or ""}</td><td class="r">{_money(d.importe) if d.importe else ""}</td><td>{E("; ".join(d.notas)[:120])}</td></tr>' for d in docs) + "</table></div></div></section>")
    parts.append('<section><div class="wrap"><h2>Cuadre de la liquidación</h2><div class="tw"><table><tr><th>Verificación</th><th class="r">Esperado</th><th class="r">Obtenido</th><th>Resultado</th></tr>' + "".join(f'<tr><td>{E(c.nombre)}</td><td class="r">{_money(c.esperado)}</td><td class="r">{_money(c.obtenido)}</td><td class="{"ok" if c.ok else "fail"}">{"OK" if c.ok else "FALLA " + _money(c.diff)}</td></tr>' for c in liq.checks) + "</table></div></div></section>")
    parts.append(f'<footer><div class="wrap"><p>Generado el {date.today():%d-%m-%Y} con Consorcio Transparente. Los hallazgos son hechos documentados que requieren aclaración de la administración; no constituyen una imputación.</p></div></footer>')
    open(out, "w", encoding="utf-8").write("<!doctype html>\n<html lang=\"es\">\n" + "\n".join(parts) + "\n</html>\n")
    return out
