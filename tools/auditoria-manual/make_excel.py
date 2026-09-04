import os
HERE = os.path.dirname(os.path.abspath(__file__))
DATOS = os.path.join(HERE, "datos") + "/"
PRIVADO = os.environ.get("CT_PRIVADO", os.path.expanduser("~/consorcio-transparente-privado")) + "/"
import json
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from alerts import ALERTS, SEV_ORDER, BULLETS, DOC_FINDINGS

SC = DATOS
D = json.load(open(SC + "data.json"))
OUT = os.path.join(HERE, "salida", "Rivadavia 2069 - Analisis expensas Agosto 2026.xlsx"); os.makedirs(os.path.dirname(OUT), exist_ok=True)

MONEY = '#,##0.00;[Red]-#,##0.00'
PCT = '0.0%'
NAVY = "1F2A44"; INK = "17253A"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
TITLE_FONT = Font(bold=True, size=15, color=INK, name="Calibri")
SUB_FONT = Font(bold=True, size=11, color=INK, name="Calibri")
NOTE_FONT = Font(italic=True, size=9, color="5A6270", name="Calibri")
BOLD = Font(bold=True, name="Calibri", size=10)
BASE = Font(name="Calibri", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="E8ECF3")
SEV_FILL = {"CRÍTICO": PatternFill("solid", fgColor="F8D0D0"), "ALTO": PatternFill("solid", fgColor="FBE3CC"),
            "MEDIO": PatternFill("solid", fgColor="FFF2C2"), "BAJO": PatternFill("solid", fgColor="E3EEF9")}
THIN = Side(style="thin", color="C9CFD9")
BORDER = Border(bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

def sheet(title):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    return ws

def title(ws, text, sub=None, row=1):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    if sub: ws.cell(row=row+1, column=1, value=sub).font = NOTE_FONT
    return row + (3 if sub else 2)

def header(ws, row, cols, widths=None):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30
    if widths:
        for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w
    return row + 1

def put(ws, row, values, fmts=None, bold=False, fill=None, wrap=False):
    for j, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = BOLD if bold else BASE
        if fmts and j - 1 < len(fmts) and fmts[j-1]: cell.number_format = fmts[j-1]
        if fill: cell.fill = fill
        if wrap: cell.alignment = WRAP
        cell.border = BORDER
    return row + 1

def total_row(ws, row, label, sums, fmts, first_col=1):
    """sums: dict col_index(1-based) -> (start_row, end_row)"""
    ws.cell(row=row, column=first_col, value=label).font = BOLD
    for c, (r0, r1) in sums.items():
        L = get_column_letter(c)
        cell = ws.cell(row=row, column=c, value=f"=SUM({L}{r0}:{L}{r1})")
        cell.font = BOLD; cell.number_format = fmts.get(c, MONEY)
    for c in range(1, ws.max_column + 1): ws.cell(row=row, column=c).fill = TOTAL_FILL
    return row + 1

def autofilter(ws, hdr_row, ncols, last_row):
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(ncols)}{last_row}"
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

G, GJ, U, EV, EF, EFJ, CO, COJ, PA, PAJ, OB, PR, SU = (D[k] for k in
  ("gastos","gastos_jul","units","evolucion","estado_fin","estado_fin_jul","composicion","composicion_jul","patrimonial","patrimonial_jul","obras","proveedores","sueldos"))
TOT_A = sum(g["importe"] for g in G); TOT_J = sum(g["importe"] for g in GJ)

# ======================================================================= 1. RESUMEN
ws = wb.active; ws.title = "Resumen"; ws.sheet_view.showGridLines = False
r = title(ws, "Consorcio Rivadavia 2069 - Análisis de la liquidación de expensas",
          f"Agosto 2026 contrastado con Julio 2026. Fuentes: liquidaciones de Administración Almazare y los 150 comprobantes adjuntos en Redconar. Elaborado 03-09-2026. Todos los importes en pesos.")
ws.column_dimensions["A"].width = 46; ws.column_dimensions["B"].width = 20; ws.column_dimensions["C"].width = 20; ws.column_dimensions["D"].width = 14; ws.column_dimensions["E"].width = 70
ws.cell(row=r, column=1, value="Indicadores clave").font = SUB_FONT; r += 1
r = header(ws, r, ["Indicador", "Agosto 2026", "Julio 2026", "Variación", "Lectura"])
deuda_a = sum(u["deuda"] for u in U if u["deuda"] > 0); deuda_j = 4283350.11
morosos_a = sum(1 for u in U if u["deuda"] > 0 and u["uf"] != 27); sinpago = sum(1 for u in U if u["deuda"] > 0 and u["pagos"] == 0)
kpis = [
 ("Gastos del mes", TOT_A, TOT_J, "Baja 8% porque julio incluyó aguinaldo y el anticipo de obra de $5 M."),
 ("Importe prorrateado a cobrar", 31705960.60, 31122626.83, "Agosto prorratea $1,83 M más que el gasto, sin concepto informado."),
 ("Expensas cobradas en el mes", 30894682.83, 21466527.20, "Agosto cobró 89,8% del saldo anterior de los propietarios."),
 ("Cobrado en término", EF["ing_termino"], EFJ["ing_termino"], ""),
 ("Cobrado de expensas adeudadas", EF["ing_adeudadas"], EFJ["ing_adeudadas"], ""),
 ("Saldo de disponibilidades al cierre", EF["saldo_cierre"], EFJ["saldo_cierre"], "Fin de junio: $12.000.224. Se consumieron $11 M en julio."),
 ("   de los cuales en efectivo (caja)", CO[1]["saldo_cierre"], COJ[1]["saldo_cierre"], "68% de la liquidez está en efectivo en poder de la administración."),
 ("   de los cuales en banco", CO[0]["saldo_cierre"], COJ[0]["saldo_cierre"], ""),
 ("Facturas pendientes de pago", -PA["facturas_pend"], -PAJ["facturas_pend"], "Incluye $140.000 no identificados en ambos meses."),
 ("Disponibilidades menos facturas pendientes", EF["saldo_cierre"] + PA["facturas_pend"], EFJ["saldo_cierre"] + PAJ["facturas_pend"], "Negativo en ambos meses."),
 ("Deuda de propietarios (bruta)", deuda_a, deuda_j, "9 unidades en agosto (11 en julio). UC-1 concentra 35%."),
 ("Intereses devengados a deudores", 535506.21, None, ""),
 ("Unidades morosas (sin contar débito por llavero)", morosos_a, 11, ""),
 ("Unidades sin ningún pago en el mes", sinpago, None, ""),
 ("Pagado en efectivo a proveedores", sum(g["importe"] for g in G if g["forma"].startswith("Efectivo")), sum(g["importe"] for g in GJ if g["forma"].startswith("Efectivo")), "Julio: seguridad $2,7 M, porcelanato $2 M, colocación $2 M."),
 ("Obras en unidades privadas (rubro + serpentinas)", 7585333.33 + 2650000, 10203457.01 + 2650000, "33% del gasto de los dos meses."),
 ("Gasto promedio por departamento (expensa del mes)", sum(u["total_mes"] for u in U if u["tipo"]=="Departamento")/94, None, "94 departamentos + 1 local + 21 cocheras."),
 ("Expensa de una cochera", 114072.19, None, ""),
]
k0 = r
for lab, a, j, note in kpis:
    ws.cell(row=r, column=1, value=lab).font = BASE
    ca = ws.cell(row=r, column=2, value=a); ca.number_format = MONEY if isinstance(a, float) else '0'; ca.font = BASE
    if j is not None:
        cj = ws.cell(row=r, column=3, value=j); cj.number_format = MONEY if isinstance(j, float) else '0'; cj.font = BASE
        if isinstance(a, float) and j:
            cv = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/C{r}-1,\"\")"); cv.number_format = PCT; cv.font = BASE
    ws.cell(row=r, column=5, value=note).font = NOTE_FONT
    for c in range(1, 6): ws.cell(row=r, column=c).border = BORDER
    r += 1
r += 1
ws.cell(row=r, column=1, value="Dónde se aplicaron los gastos mayores (julio + agosto = $62.420.048)").font = SUB_FONT; r += 1
r = header(ws, r, ["Destino", "Importe 2 meses", "% del gasto 2 meses", "", "Detalle"])
b0 = r
for lab, amt, det in BULLETS:
    ws.cell(row=r, column=1, value=lab).font = BASE
    c = ws.cell(row=r, column=2, value=amt); c.number_format = MONEY; c.font = BASE
    c = ws.cell(row=r, column=3, value=f"=B{r}/{TOT_A + TOT_J:.2f}"); c.number_format = PCT; c.font = BASE
    ws.cell(row=r, column=5, value=det).font = BASE; ws.cell(row=r, column=5).alignment = WRAP
    for cc in range(1, 6): ws.cell(row=r, column=cc).border = BORDER
    ws.row_dimensions[r].height = 30
    r += 1
r += 1
ws.cell(row=r, column=1, value="Problemas detectados por severidad").font = SUB_FONT; r += 1
r = header(ws, r, ["Severidad", "Cantidad", "", "", "Ver hoja 'Problemas detectados'"])
for sev in ("CRÍTICO","ALTO","MEDIO","BAJO"):
    ws.cell(row=r, column=1, value=sev).fill = SEV_FILL[sev]; ws.cell(row=r, column=1).font = BOLD
    ws.cell(row=r, column=2, value=sum(1 for a in ALERTS if a[0]==sev)).font = BASE
    r += 1
r += 1
ws.cell(row=r, column=1, value="Estado financiero (Ley 941 art. 10 inc. c)").font = SUB_FONT; r += 1
r = header(ws, r, ["Concepto", "Agosto 2026", "Julio 2026", "", ""])
for lab, k in [("Saldo anterior","saldo_anterior"),("Ingresos por expensas en término","ing_termino"),("Ingresos por expensas adeudadas","ing_adeudadas"),
               ("Ingresos por intereses","ing_intereses"),("Ingresos por expensas adelantadas","ing_adelantadas"),("Egresos por gastos del mes","egresos"),("Saldo al cierre","saldo_cierre")]:
    a = EF[k]; j = EFJ[k]
    if k == "egresos": a, j = -a, -j
    r = put(ws, r, [lab, a, j], [None, MONEY, MONEY], bold=(k in ("saldo_cierre",)))
r += 1
ws.cell(row=r, column=1, value="Composición de las disponibilidades").font = SUB_FONT; r += 1
r = header(ws, r, ["Cuenta / mes", "Saldo anterior", "Ingresos", "Egresos", "Saldo al cierre"])
ws.column_dimensions["D"].width = 18
for lab, rows in (("Agosto 2026", CO), ("Julio 2026", COJ)):
    for c in rows:
        r = put(ws, r, [f"{lab} - {c['cuenta']}", c["saldo_ant"], c["ingresos"], -c["egresos"], c["saldo_cierre"]], [None, MONEY, MONEY, MONEY, MONEY])
r += 1
ws.cell(row=r, column=1, value="Estado patrimonial").font = SUB_FONT; r += 1
r = header(ws, r, ["Concepto", "Agosto 2026", "Julio 2026", "", ""])
for lab, k in [("Disponibilidades al cierre","disponibilidades"),("Expensas y otros conceptos a cobrar","a_cobrar"),("Facturas pendientes de pago","facturas_pend"),("Total patrimonial","total")]:
    r = put(ws, r, [lab, PA[k], PAJ[k]], [None, MONEY, MONEY], bold=(k=="total"))
r += 1
ws.cell(row=r, column=1, value="Referencia interanual: Septiembre 2025").font = SUB_FONT; r += 1
rs = D["ref_sep25"]
r = header(ws, r, ["Concepto", "Sept 2025", "Agosto 2026", "Variación", ""])
for lab, a, b in [("Gastos del mes", rs["gastos"], TOT_A), ("Sueldos y cargas", rs["sueldos"], 5945356.92), ("Mantenimientos en unidades", rs["mant_unidades"], 7585333.33),
                  ("Metrogas cuenta principal (bimestre 04)", rs["metrogas_principal"], 3482893.85), ("Disponibilidades", rs["disponibilidades"], EF["saldo_cierre"]), ("   en caja (efectivo)", rs["caja"], CO[1]["saldo_cierre"])]:
    ws.cell(row=r, column=1, value=lab).font = BASE
    ws.cell(row=r, column=2, value=a).number_format = MONEY; ws.cell(row=r, column=3, value=b).number_format = MONEY
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = PCT
    for c in range(1, 5): ws.cell(row=r, column=c).border = BORDER
    r += 1

# ======================================================================= 2. PROBLEMAS
ws = sheet("Problemas detectados")
r = title(ws, "Problemas detectados", "Ordenados por severidad. 'Monto' es el importe involucrado o afectado, no una pérdida estimada.")
r = header(ws, r, ["#", "Severidad", "Área", "Problema", "Evidencia", "Monto involucrado", "Recomendación"], [4, 11, 22, 45, 80, 18, 55])
for i, (sev, area, tit, ev, monto, rec) in enumerate(sorted(ALERTS, key=lambda a: SEV_ORDER[a[0]]), 1):
    put(ws, r, [i, sev, area, tit, ev, monto, rec], [None, None, None, None, None, MONEY, None], wrap=True)
    ws.cell(row=r, column=2).fill = SEV_FILL[sev]; ws.cell(row=r, column=2).font = BOLD; ws.cell(row=r, column=4).font = BOLD
    ws.row_dimensions[r].height = 110
    r += 1
autofilter(ws, 4, 7, r - 1)

# ======================================================================= 3. GASTOS POR RUBRO
ws = sheet("Gastos por rubro")
r = title(ws, "Gastos por rubro: Agosto vs. Julio 2026", "Las dos categorías 'Servicios públicos' del PDF se unifican. Columnas A/B/D = clases de prorrateo (A fijo, B variable, D gas y calefacción sin cocheras).")
cats = OrderedDict()
for g in G: cats.setdefault(g["categoria"], [0,0,0,0,0])
for g in GJ: cats.setdefault(g["categoria"], [0,0,0,0,0])
for g in G:
    cats[g["categoria"]][0] += g["importe"]; cats[g["categoria"]][{"A":2,"B":3,"D":4}[g["columna"]]] += g["importe"]
for g in GJ: cats[g["categoria"]][1] += g["importe"]
r = header(ws, r, ["Rubro", "Agosto 2026", "% agosto", "Julio 2026", "% julio", "Variación", "Ago col. A", "Ago col. B", "Ago col. D", "Líneas ago"], [34, 18, 10, 18, 10, 11, 16, 14, 16, 10])
r0 = r
for cat, (a, j, ca, cb, cd) in sorted(cats.items(), key=lambda x: -x[1][0]):
    ws.cell(row=r, column=1, value=cat).font = BASE
    ws.cell(row=r, column=2, value=round(a,2)).number_format = MONEY
    ws.cell(row=r, column=3, value=f"=B{r}/B${{T}}").number_format = PCT
    ws.cell(row=r, column=4, value=round(j,2)).number_format = MONEY
    ws.cell(row=r, column=5, value=f"=D{r}/D${{T}}").number_format = PCT
    ws.cell(row=r, column=6, value=f"=IFERROR(B{r}/D{r}-1,\"\")").number_format = PCT
    ws.cell(row=r, column=7, value=round(ca,2)).number_format = MONEY; ws.cell(row=r, column=8, value=round(cb,2)).number_format = MONEY; ws.cell(row=r, column=9, value=round(cd,2)).number_format = MONEY
    ws.cell(row=r, column=10, value=sum(1 for g in G if g["categoria"]==cat))
    for c in range(1, 11): ws.cell(row=r, column=c).border = BORDER
    r += 1
T = r
for rr in range(r0, r):
    for c in (3, 5): ws.cell(row=rr, column=c).value = ws.cell(row=rr, column=c).value.replace("{T}", str(T))
total_row(ws, r, "Total", {2:(r0,r-1), 4:(r0,r-1), 7:(r0,r-1), 8:(r0,r-1), 9:(r0,r-1), 10:(r0,r-1)}, {10:'0'})
ws.cell(row=r, column=3, value=f"=B{r}/B{T}").number_format = PCT; ws.cell(row=r, column=5, value=f"=D{r}/D{T}").number_format = PCT
ws.cell(row=r, column=6, value=f"=B{r}/D{r}-1").number_format = PCT
ws.freeze_panes = "A5"

# ======================================================================= 4/5. DETALLE DE GASTOS
def detalle(name, rows, sub, full=True):
    ws = sheet(name)
    r = title(ws, name, sub)
    cols = ["#", "Rubro", "Proveedor", "Concepto", "Col.", "Importe pagado", "Fecha factura", "N° factura", "Importe factura", "Fecha pago", "Forma de pago", "Período", "Días factura→pago", "% del mes"] if full else \
           ["#", "Rubro", "Proveedor", "Concepto", "Col.", "Importe pagado", "Fecha factura", "Fecha pago", "Forma de pago", "Días factura→pago", "% del mes"]
    widths = [4, 26, 30, 70, 5, 16, 12, 18, 16, 12, 18, 14, 10, 9] if full else [4, 26, 30, 70, 5, 16, 12, 12, 18, 10, 9]
    r = header(ws, r, cols, widths); r0 = r
    tot = sum(g["importe"] for g in rows)
    for g in rows:
        if full:
            vals = [g["n"], g["categoria"], g["proveedor"], g["concepto"], g["columna"], g["importe"], g["fecha_factura"], g["nro_factura"], g["importe_factura"], g["fecha_pago"], g["forma"], g["periodo"], g["dias_factura_pago"], None]
            fm = [None]*5 + [MONEY, None, None, MONEY, None, None, None, '0', PCT]
        else:
            vals = [g["n"], g["categoria"], g["proveedor"], g["concepto"], g["columna"], g["importe"], g["fecha_factura"], g["fecha_pago"], g["forma"], g["dias_factura_pago"], None]
            fm = [None]*5 + [MONEY, None, None, None, '0', PCT]
        put(ws, r, vals, fm, wrap=True)
        ws.cell(row=r, column=len(cols), value=f"=F{r}/{tot:.2f}").number_format = PCT
        if g["forma"].startswith("Efectivo"): ws.cell(row=r, column=(11 if full else 9)).fill = SEV_FILL["ALTO"]
        if g["dias_factura_pago"] is not None and (g["dias_factura_pago"] > 60 or g["dias_factura_pago"] < 0): ws.cell(row=r, column=(13 if full else 10)).fill = SEV_FILL["MEDIO"]
        ws.row_dimensions[r].height = 42
        r += 1
    total_row(ws, r, "Total", {6:(r0,r-1)}, {})
    if full: ws.cell(row=r, column=9, value=f"=SUM(I{r0}:I{r-1})").number_format = MONEY
    autofilter(ws, 4, len(cols), r - 1)
    return ws
detalle("Gastos Agosto detalle", G, "Las 43 líneas de la liquidación de agosto. Naranja = pagado en efectivo; amarillo = más de 60 días entre factura y pago, o pago anterior a la factura.")
detalle("Gastos Julio detalle", GJ, "Las 44 líneas de la liquidación de julio (mes de contraste). Naranja = pagado en efectivo; amarillo = atraso o pago anterior a la factura.", full=False)

# ======================================================================= 6. PROVEEDORES
ws = sheet("Proveedores")
r = title(ws, "Ranking de proveedores: quiénes se llevan la plata (julio + agosto 2026)", "Los sueldos figuran como 'Consorcio Rivadavia 2069 (sueldos)' y ARCA incluye cargas sociales y retenciones. Ordenado por total de dos meses.")
prov = OrderedDict()
for g in G: prov.setdefault(g["proveedor"], [0,0,0]); prov[g["proveedor"]][0] += g["importe"]; prov[g["proveedor"]][2] += 1
for g in GJ: prov.setdefault(g["proveedor"], [0,0,0]); prov[g["proveedor"]][1] += g["importe"]; prov[g["proveedor"]][2] += 1
info = {p["fantasia"]: p for p in PR}
r = header(ws, r, ["#", "Proveedor", "Razón social", "CUIT", "Agosto 2026", "Julio 2026", "Total 2 meses", "% del gasto 2 meses", "% acumulado", "Líneas", "Rubro principal"], [4, 34, 40, 15, 16, 16, 16, 12, 12, 8, 30])
r0 = r; TOT2 = TOT_A + TOT_J
for i, (p, (a, j, n)) in enumerate(sorted(prov.items(), key=lambda x: -(x[1][0]+x[1][1])), 1):
    inf = info.get(p, {}); cat = next((g["categoria"] for g in G + GJ if g["proveedor"] == p), "")
    put(ws, r, [i, p, inf.get("razon_social", ""), inf.get("cuit", ""), round(a,2), round(j,2), None, None, None, n, cat], [None]*4 + [MONEY, MONEY, MONEY, PCT, PCT, '0', None])
    ws.cell(row=r, column=7, value=f"=E{r}+F{r}").number_format = MONEY
    ws.cell(row=r, column=8, value=f"=G{r}/{TOT2:.2f}").number_format = PCT
    ws.cell(row=r, column=9, value=f"=SUM(H${r0}:H{r})").number_format = PCT
    r += 1
total_row(ws, r, "Total", {5:(r0,r-1), 6:(r0,r-1), 7:(r0,r-1), 8:(r0,r-1)}, {8: PCT})
autofilter(ws, 4, 11, r - 1)

# ======================================================================= 7. DEUDORES
ws = sheet("Deudores")
r = title(ws, "Propietarios con saldo deudor: Agosto 2026 vs. Julio 2026", "Meses de deuda = deuda / expensa mensual de la unidad. Tasa = interés del mes / deuda. La UF 27 no es mora: es un débito por llavero.")
dj = {x["uf"]: x["deuda"] for x in D["deudores_jul"]}
r = header(ws, r, ["UF", "Piso-Depto", "Propietario", "Tipo", "Deuda agosto", "Deuda julio", "Variación", "Pagó en agosto", "Interés del mes", "Tasa s/ deuda", "Expensa mensual", "Meses de deuda", "% de la deuda total", "% acumulado", "Estado"],
           [6, 10, 28, 13, 16, 16, 14, 16, 14, 10, 15, 10, 12, 11, 34])
r0 = r
deudores = sorted([u for u in U if u["deuda"] > 0], key=lambda u: -u["deuda"])
for u in deudores:
    put(ws, r, [u["uf"], u["piso_depto"], u["propietario"], u["tipo"], u["deuda"], dj.get(u["uf"]), None, u["pagos"], u["interes"], u["tasa_int_sobre_deuda"], u["total_mes"], u["meses_deuda"], None, None, u["estado"]],
        [None]*4 + [MONEY, MONEY, MONEY, MONEY, MONEY, PCT, MONEY, '0.0', PCT, PCT, None])
    ws.cell(row=r, column=7, value=f"=IF(F{r}=\"\",E{r},E{r}-F{r})").number_format = MONEY
    ws.cell(row=r, column=13, value=f"=E{r}/{deuda_a:.2f}").number_format = PCT
    ws.cell(row=r, column=14, value=f"=SUM(M${r0}:M{r})").number_format = PCT
    if u["pagos"] == 0: ws.cell(row=r, column=8).fill = SEV_FILL["CRÍTICO"]
    if u["meses_deuda"] >= 3: ws.cell(row=r, column=12).fill = SEV_FILL["CRÍTICO"]
    r += 1
total_row(ws, r, "Total agosto", {5:(r0,r-1), 8:(r0,r-1), 9:(r0,r-1), 13:(r0,r-1)}, {13: PCT})
ws.cell(row=r, column=6, value=deuda_j).number_format = MONEY; ws.cell(row=r, column=6).font = BOLD
r += 2
ws.cell(row=r, column=1, value="Deudores de julio que regularizaron en agosto").font = SUB_FONT; r += 1
r = header(ws, r, ["UF", "Piso-Depto", "Propietario", "", "Deuda julio", "Pagó en agosto", "Estado agosto"])
ua = {u["uf"]: u for u in U}
for x in D["deudores_jul"]:
    if ua[x["uf"]]["deuda"] <= 0:
        r = put(ws, r, [x["uf"], x["piso_depto"], x["propietario"], "", x["deuda"], ua[x["uf"]]["pagos"], ua[x["uf"]]["estado"]], [None]*4 + [MONEY, MONEY, None])
r += 1
ws.cell(row=r, column=1, value="Nuevos deudores en agosto").font = SUB_FONT; r += 1
r = header(ws, r, ["UF", "Piso-Depto", "Propietario", "", "Deuda agosto", "Pagó en agosto", "Estado agosto"])
for u in deudores:
    if u["uf"] not in dj:
        r = put(ws, r, [u["uf"], u["piso_depto"], u["propietario"], "", u["deuda"], u["pagos"], u["estado"]], [None]*4 + [MONEY, MONEY, None])
ws.freeze_panes = "A5"

# ======================================================================= 8. GASTO POR UNIDAD FUNCIONAL
ws = sheet("Gasto por UF")
r = title(ws, "¿Dónde se gastó la plata? Obras y gastos aplicados a unidades funcionales concretas (julio + agosto 2026)",
          "Para las serpentinas de Roth (3 cuotas de $2.650.000) se prorratea cada cuota entre 13-B, 12-B y la cupla según el valor de cada trabajo.")
r = header(ws, r, ["UF beneficiaria", "Propietario", "Obra / concepto", "Proveedor", "Costo total", "Pagado julio", "Pagado agosto", "Pagado 2 meses", "Pendiente", "Columna de prorrateo", "Observación"],
           [30, 34, 48, 30, 16, 16, 16, 16, 16, 18, 60])
r0 = r
for o in OB:
    put(ws, r, [o["uf"], o["propietario"], o["obra"], o["proveedor"], o["total_obra"], o["pagado_jul"], o["pagado_ago"], None, None, o["columna"], o["obs"]], [None]*4 + [MONEY]*5 + [None, None], wrap=True)
    ws.cell(row=r, column=8, value=f"=F{r}+G{r}").number_format = MONEY; ws.cell(row=r, column=9, value=f"=E{r}-H{r}").number_format = MONEY
    ws.row_dimensions[r].height = 40; r += 1
total_row(ws, r, "Total", {5:(r0,r-1), 6:(r0,r-1), 7:(r0,r-1), 8:(r0,r-1), 9:(r0,r-1)}, {})
r += 2
ws.cell(row=r, column=1, value="Resumen por unidad beneficiaria").font = SUB_FONT; r += 1
r = header(ws, r, ["UF beneficiaria", "Propietario", "Cantidad de trabajos", "", "Costo total", "Pagado julio", "Pagado agosto", "Pagado 2 meses", "Pendiente", "% del gasto 2 meses"])
agg = OrderedDict()
for o in OB:
    a = agg.setdefault(o["uf"], [o["propietario"], 0, 0, 0, 0]); a[1] += 1; a[2] += o["total_obra"]; a[3] += o["pagado_jul"]; a[4] += o["pagado_ago"]
r1 = r
for uf, (prop, n, tot, pj, pa) in sorted(agg.items(), key=lambda x: -x[1][2]):
    put(ws, r, [uf, prop, n, "", round(tot,2), round(pj,2), round(pa,2), None, None, None], [None, None, '0', None, MONEY, MONEY, MONEY, MONEY, MONEY, PCT])
    ws.cell(row=r, column=8, value=f"=F{r}+G{r}").number_format = MONEY; ws.cell(row=r, column=9, value=f"=E{r}-H{r}").number_format = MONEY
    ws.cell(row=r, column=10, value=f"=H{r}/{TOT2:.2f}").number_format = PCT
    r += 1
total_row(ws, r, "Total", {5:(r1,r-1), 6:(r1,r-1), 7:(r1,r-1), 8:(r1,r-1), 9:(r1,r-1), 10:(r1,r-1)}, {10: PCT})
ws.freeze_panes = "A5"

# ======================================================================= 9. ESTADO DE CUENTAS
ws = sheet("Estado de cuentas")
r = title(ws, "Estado de cuentas y prorrateo por unidad: Agosto 2026 (116 unidades)", "Reproducción del cuadro del PDF con columnas calculadas. Deuda = saldo anterior − pagos + créditos/débitos. A pagar = expensa del mes + deuda + interés + redondeo.")
r = header(ws, r, ["UF", "Piso-Depto", "Propietario", "Tipo", "Saldo anterior", "Pagos", "Créd./Déb.", "Deuda", "Interés", "% A", "Expensa A", "% B", "Expensa B", "% D", "Expensa D", "Expensa del mes", "Redondeo", "A pagar", "Pagó % del saldo", "Meses de deuda", "Estado"],
           [6, 10, 30, 13, 15, 15, 12, 14, 12, 7, 14, 7, 12, 7, 14, 15, 9, 15, 10, 9, 36])
r0 = r
for u in U:
    put(ws, r, [u["uf"], u["piso_depto"], u["propietario"], u["tipo"], u["saldo_ant"], u["pagos"], u["cred_deb"], u["deuda"], u["interes"], u["pct_A"]/100, u["exp_A"], u["pct_B"]/100, u["exp_B"], u["pct_D"]/100, u["exp_D"], u["total_mes"], u["redondeo"], u["a_pagar"], u["pago_pct_saldo"], u["meses_deuda"], u["estado"]],
        [None]*4 + [MONEY, MONEY, MONEY, MONEY, MONEY, '0.00%', MONEY, '0.00%', MONEY, '0.00%', MONEY, MONEY, MONEY, MONEY, PCT, '0.0', None])
    if u["deuda"] > 0 and u["uf"] != 27: ws.cell(row=r, column=8).fill = SEV_FILL["CRÍTICO"] if u["pagos"] == 0 else SEV_FILL["ALTO"]
    if u["deuda"] < 0: ws.cell(row=r, column=8).fill = SEV_FILL["BAJO"]
    if u["interes"] > 0: ws.cell(row=r, column=9).fill = SEV_FILL["MEDIO"]
    r += 1
total_row(ws, r, "Total", {c:(r0,r-1) for c in (5,6,7,8,9,10,11,12,13,14,15,16,17,18)}, {10:'0.00%', 12:'0.00%', 14:'0.00%'})
autofilter(ws, 4, 21, r - 1)

# ======================================================================= 10. RECARGOS Y SALDOS A FAVOR
ws = sheet("Recargos y saldos a favor")
r = title(ws, "Unidades con interés/recargo, crédito o saldo a favor en agosto", "Recargo del 5% = pagó después del 1° vencimiento. Saldo a favor = pagó de más o adelantó expensas.")
r = header(ws, r, ["UF", "Piso-Depto", "Propietario", "Saldo anterior", "Pagos", "Créd./Déb.", "Deuda (neg. = a favor)", "Interés / recargo", "Recargo s/ saldo anterior", "Estado"], [6, 10, 30, 15, 15, 12, 16, 14, 12, 40])
for u in U:
    if u["interes"] > 0 or u["deuda"] < 0 or u["cred_deb"] != 0:
        r = put(ws, r, [u["uf"], u["piso_depto"], u["propietario"], u["saldo_ant"], u["pagos"], u["cred_deb"], u["deuda"], u["interes"], u["tasa_int_sobre_saldo_ant"], u["estado"]], [None]*3 + [MONEY]*5 + [PCT, None])
ws.freeze_panes = "A5"

# ======================================================================= 11. EVOLUCIÓN
ws = sheet("Evolución")
r = title(ws, "Evolución mensual: prorrateado, gastado y cobrado (febrero → agosto 2026)", "Febrero proviene de la liquidación de julio; el resto coincide en ambas liquidaciones. Saldos de cierre solo para los meses con liquidación disponible.")
ev = [dict(mes="Febrero 2026", a_cobrar=21134517.12, gastos=19521427.25, cobrado=18592951.79)] + EV
r = header(ws, r, ["Mes", "Importe a cobrar (prorrateado)", "Gastos del mes", "Expensas cobradas", "Cobrado / prorrateado", "Gastos / prorrateado", "Prorrateado − gastos", "Cobrado − gastos", "Saldo disponibilidades al cierre"], [16, 20, 18, 18, 14, 14, 18, 18, 20])
r0 = r
saldos = {"Junio 2026": 12000224.34, "Julio 2026": 923626.64, "Agosto 2026": 1941386.31}
for e in ev:
    put(ws, r, [e["mes"], e["a_cobrar"], e["gastos"], e["cobrado"], None, None, None, None, saldos.get(e["mes"])], [None, MONEY, MONEY, MONEY, PCT, PCT, MONEY, MONEY, MONEY])
    ws.cell(row=r, column=5, value=f"=D{r}/B{r}").number_format = PCT; ws.cell(row=r, column=6, value=f"=C{r}/B{r}").number_format = PCT
    ws.cell(row=r, column=7, value=f"=B{r}-C{r}").number_format = MONEY; ws.cell(row=r, column=8, value=f"=D{r}-C{r}").number_format = MONEY
    r += 1
total_row(ws, r, "Total 7 meses", {2:(r0,r-1), 3:(r0,r-1), 4:(r0,r-1), 7:(r0,r-1), 8:(r0,r-1)}, {})
ws.cell(row=r, column=5, value=f"=D{r}/B{r}").number_format = PCT; ws.cell(row=r, column=6, value=f"=C{r}/B{r}").number_format = PCT
r += 2
ws.cell(row=r, column=1, value="Referencia Septiembre 2025 (liquidación anterior disponible)").font = SUB_FONT; r += 1
r = put(ws, r, ["Septiembre 2025", 18157655.66, 20400218.20, 17743553.93, None, None, None, None, 7294814.32], [None, MONEY, MONEY, MONEY, PCT, PCT, MONEY, MONEY, MONEY])
ws.cell(row=r-1, column=5, value=f"=D{r-1}/B{r-1}").number_format = PCT; ws.cell(row=r-1, column=6, value=f"=C{r-1}/B{r-1}").number_format = PCT
ws.cell(row=r-1, column=7, value=f"=B{r-1}-C{r-1}").number_format = MONEY; ws.cell(row=r-1, column=8, value=f"=D{r-1}-C{r-1}").number_format = MONEY

# ======================================================================= 12. SUELDOS
ws = sheet("Anexo sueldos")
r = title(ws, "Detalle de sueldos y cargas sociales: período Julio 2026 (pagado en agosto)", "Recibos 584766 y 584774. Cargas patronales (F.931) del período se pagan al mes siguiente.")
ws.column_dimensions["A"].width = 46; ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 16; ws.column_dimensions["D"].width = 16; ws.column_dimensions["E"].width = 16
for s in SU:
    ws.cell(row=r, column=1, value=f"{s['empleado']}  -  {s['cargo']}  -  CUIL {s['cuil']}").font = SUB_FONT; r += 1
    r = header(ws, r, ["Concepto", "Cantidad", "Haberes", "Deducciones", "Neto"])
    r0 = r
    for c, q, h, dd in s["items"]:
        r = put(ws, r, [c, q, h or None, dd or None], [None, '0.00##', MONEY, MONEY])
    total_row(ws, r, "Total", {3:(r0,r-1), 4:(r0,r-1)}, {})
    ws.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = MONEY; ws.cell(row=r, column=5).font = BOLD
    r += 2
ws.cell(row=r, column=1, value="Costo laboral del mes (agosto) y comparación").font = SUB_FONT; r += 1
r = header(ws, r, ["Concepto", "", "Agosto 2026", "Julio 2026", "Nota"])
for lab, a, j, n in [("Sueldos netos", 1424799+1774010, 2370332+2750537, "Julio incluye SAC"), ("F.931 (período anterior)", 1920211.16, 3066158.25, "Pagados con un mes de atraso"),
                     ("FATERYH + SUTERH + SERACARH", 365166.81+184418.86+20490.98, 570647.58+294490.78+32721.20, ""), ("Retención SIRE C.S.I. (mal clasificada aquí)", 256260.11, 255132.48, "Corresponde al servicio de seguridad"),
                     ("Total rubro según liquidación", 5945356.92, 9340019.29, ""), ("Seguro de vida Berkley", 39711.71, 39711.71, "Sumas aseguradas desactualizadas")]:
    r = put(ws, r, [lab, "", a, j, n], [None, None, MONEY, MONEY, None])

# ======================================================================= 13. OBRAS EN CUOTAS
ws = sheet("Obras y compromisos")
r = title(ws, "Obras, compromisos en cuotas y facturas pendientes", "Seguimiento de lo que el consorcio ya comprometió y todavía debe pagar.")
r = header(ws, r, ["Proveedor", "Obra / contrato", "Total contrato", "Pagado antes de julio", "Pagado julio", "Pagado agosto", "Pagado acumulado", "Pendiente", "Cuotas / estado", "Observación"], [30, 50, 16, 16, 16, 16, 16, 16, 22, 60])
r0 = r
for row in [
 ("PEÑALOZA ALEJANDRO ROBERTO", "Columna calefacción sector F pisos 8-12", 14000000, 0, 5000000, 4333333.33, "2 cuotas de $2.333.333,33", "Factura 201 (jul) emitida después del pago; factura 202 (ago) por $9 M"),
 ("MARIO LEONARDO ROTH", "Serpentinas 12-B / 13-B + cupla", 7950000, 2650000, 2650000, 2650000, "3/3 pagadas (1ª presumiblemente junio)", "Mismo proveedor certifica los equipos térmicos"),
 ("SACZEWICZYK MARIA EUGENIA", "Porcelanato 13-B (colocación)", 4552000, 0, 2000000, 2552000, "Pagada", "$2 M en efectivo"),
 ("SACZEWICZYK MARIA EUGENIA", "Pintura 13-B", 700000, 0, 0, 700000, "Pagada", "Factura y pago el 31-08"),
 ("LO & CO S.A.", "Porcelanato 13-B (material)", 2003457.01, 0, 2003457.01, 0, "Pagada", "En efectivo"),
 ("ALLIANZ SEGUROS", "Póliza integral (10 cuotas)", 2686520, 268652*6, 268652, 268652, "8/10", ""),
 ("ALLIANZ SEGUROS", "Endoso 2 RC $650,1 M (6 cuotas)", 2294262, 382377*2, 382377, 382377, "4/6", "Cuesta más que la póliza base"),
 ("ALLIANZ SEGUROS", "Endoso 1 (7 cuotas)", 56896, 8128*3, 8128, 8128, "5/7", ""),
 ("BERKLEY CIA DE SEGUROS", "Vida colectivo (10 cuotas)", 397117.10, 39711.71*5, 39711.71, 39711.71, "7/10", ""),
 ("No identificado", "Diferencia en 'Facturas pendientes de pago'", 140000, 0, 0, 0, "Pendiente en jul y ago", "No corresponde a ninguna factura informada"),
]:
    prov_, obra_, tot, pre, pj, pa, est, obs = row
    put(ws, r, [prov_, obra_, tot, pre, pj, pa, None, None, est, obs], [None, None, MONEY, MONEY, MONEY, MONEY, MONEY, MONEY, None, None], wrap=True)
    ws.cell(row=r, column=7, value=f"=D{r}+E{r}+F{r}").number_format = MONEY; ws.cell(row=r, column=8, value=f"=C{r}-G{r}").number_format = MONEY
    ws.row_dimensions[r].height = 32; r += 1
total_row(ws, r, "Total", {3:(r0,r-1), 4:(r0,r-1), 5:(r0,r-1), 6:(r0,r-1), 7:(r0,r-1), 8:(r0,r-1)}, {})
r += 2
r = put(ws, r, ["Facturas pendientes informadas por la administración al 31-08-2026", "", -PA["facturas_pend"]], [None, None, MONEY], bold=True)
r = put(ws, r, ["Disponibilidades al 31-08-2026", "", EF["saldo_cierre"]], [None, None, MONEY], bold=True)
r = put(ws, r, ["Déficit de cobertura", "", EF["saldo_cierre"] + PA["facturas_pend"]], [None, None, MONEY], bold=True)

# ======================================================================= 14. PROVEEDORES (padrón)
ws = sheet("Padrón proveedores")
r = title(ws, "Padrón de proveedores informado en la liquidación de agosto", "Datos tal como figuran en el PDF. Vacío = no informado.")
r = header(ws, r, ["Nombre de fantasía", "Razón social", "CUIT", "Dirección", "Pagado agosto", "Líneas agosto"], [34, 48, 16, 40, 16, 10])
for p in sorted(PR, key=lambda p: -p["pagado_mes"]):
    r = put(ws, r, [p["fantasia"], p["razon_social"], p["cuit"] or None, p["direccion"] or None, p["pagado_mes"], p["lineas"]], [None, None, None, None, MONEY, '0'])
    if not p["cuit"] and p["fantasia"] not in ("AGIP-ABL","ARCA","BANCO DE GALICIA","FATERYH","SERACARH","SUTERH"): ws.cell(row=r-1, column=3).fill = SEV_FILL["MEDIO"]
ws.freeze_panes = "A5"

# ======================================================================= 15. COMPROBANTES REDCONAR
ws = sheet("Hallazgos comprobantes")
r = title(ws, "Hallazgos en los comprobantes de Redconar (julio y agosto 2026)", "Se descargaron y leyeron los 150 adjuntos (facturas y comprobantes de pago) del portal de propietarios. Ordenados por severidad.")
r = header(ws, r, ["Mes", "Fecha", "Proveedor según liquidación", "Importe", "Documento", "Hallazgo", "Severidad"], [9, 12, 34, 16, 34, 90, 11])
for mes, fecha, prov, imp, doc, hall, sev in sorted(DOC_FINDINGS, key=lambda x: SEV_ORDER[x[6]]):
    put(ws, r, [mes, fecha, prov, imp, doc, hall, sev], [None, None, None, MONEY, None, None, None], wrap=True)
    ws.cell(row=r, column=7).fill = SEV_FILL[sev]; ws.cell(row=r, column=7).font = BOLD
    ws.row_dimensions[r].height = 60; r += 1
autofilter(ws, 4, 7, r - 1)

ws = sheet("Comprobantes Redconar")
r = title(ws, "Índice de los 150 comprobantes descargados de Redconar", "Carpeta: ~/Descargas/Comprobantes Rivadavia 2069/<mes>/. 'E' = adjunto del egreso (pago), 'T' = adjunto de la factura (ticket). Las líneas '(sin adjuntos)' no tienen ningún respaldo en el portal.")
manifest = json.load(open(SC + "manifest.json"))
r = header(ws, r, ["Mes", "#", "Fecha", "Proveedor (Redconar)", "Importe", "Caja / forma", "N° factura", "Rubro", "Descripción", "Tipo adjunto", "Nombre del adjunto", "Archivo"], [15, 5, 12, 34, 16, 26, 16, 26, 50, 10, 40, 70])
import re as _re
for m in manifest:
    imp = float(_re.sub(r"[^\d.]", "", m.get("valor", "").replace(",", ""))) if m.get("valor") else None
    put(ws, r, [m["mes"], m["n"], m["fecha"], m["proveedor"], imp, m.get("caja"), m.get("factura"), m.get("categoria"), m.get("desc"), m.get("src", ""), m["nombre"], m.get("archivo", "")], [None, '0', None, None, MONEY] + [None]*7)
    if m["nombre"] == "(sin adjuntos)": ws.cell(row=r, column=11).fill = SEV_FILL["MEDIO"]
    if "Efectivo" in (m.get("caja") or "") or "CAJA" in (m.get("caja") or ""): ws.cell(row=r, column=6).fill = SEV_FILL["ALTO"]
    r += 1
autofilter(ws, 4, 12, r - 1)

for w in wb.worksheets:
    w.sheet_properties.tabColor = {"Resumen": NAVY, "Problemas detectados": "C0392B", "Hallazgos comprobantes": "C0392B"}.get(w.title, "8FA3BF")
wb.save(OUT)
print("saved", OUT)
